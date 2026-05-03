"""Equivalent of the user's `DiagnoseHeyRoyaWorksheet` VBA, but reads the
.xlsx directly via openpyxl. This avoids needing to open Excel.
"""

import sys
from openpyxl import load_workbook

PATH = r"C:\Users\carin\Downloads\HeyRoya_Pro_Correction_Worksheet_Template-20260503-201810.xlsx"

print("HEYROYA WORKBOOK DIAGNOSTICS")
print("-" * 40)

# data_only=False keeps formulas as text; load() does it once
wb = load_workbook(PATH, data_only=False)

# --- MODULE SCAN ---
# .xlsx (vs .xlsm) cannot contain VBA modules — there is no vbaProject.bin
# in the archive. openpyxl exposes wb.vba_archive (will be None for .xlsx).
print("\nMODULES FOUND:")
if wb.vba_archive is None:
    print("  (none — file is .xlsx, contains no VBA project)")
else:
    print("  vbaProject.bin present")

# --- ACTIVE SHEET ---
ws = wb.active
print(f"\nActive Worksheet: {ws.title}")

# --- CHECK TIMESTAMP ---
b5 = ws["B5"]
print("\nGenerated timestamp:")
print(f"  Value: {b5.value}")
print(f"  Type:  {type(b5.value).__name__}")
print(f"  NumberFormat: {b5.number_format}")

# --- CHECK COMPLETION FORMULA ---
b6 = ws["B6"]
print("\nCompletion formula:")
print(f"  B6 Formula: {b6.value}")

# --- CHECK STATUS FORMULA ---
b7 = ws["B7"]
print("\nStatus formula:")
print(f"  B7 Formula: {b7.value}")

# --- CHECK FREEZE PANES ---
print(f"\nFreeze panes: {ws.freeze_panes}")

# --- CHECK PROTECTION ---
print(f"Worksheet protected: {ws.protection.sheet}")
print(f"Protection password set: {ws.protection.password is not None}")

# --- CHECK CONDITIONAL FORMATTING ---
print("\nConditional formatting rules on E12:E500:")
for ref, rules in ws.conditional_formatting._cf_rules.items():
    ref_str = str(ref) if not hasattr(ref, "ranges") else " ".join(str(r) for r in ref.ranges)
    if "E12:E500" in ref_str or "E12" in ref_str:
        for r in rules:
            ftype = getattr(r, "type", "?")
            formula = getattr(r, "formula", None)
            formula_str = formula[0] if formula else "(none)"
            print(f"  - Range: {ref_str} | Type: {ftype} | Formula: {formula_str}")

# Show ALL CF rules for completeness
print("\nALL conditional formatting rules in the workbook:")
for ref, rules in ws.conditional_formatting._cf_rules.items():
    ref_str = str(ref) if not hasattr(ref, "ranges") else " ".join(str(r) for r in ref.ranges)
    for r in rules:
        ftype = getattr(r, "type", "?")
        formula = getattr(r, "formula", None)
        formula_str = formula[0] if formula else "(none)"
        print(f"  - Range: {ref_str} | Type: {ftype} | Formula: {formula_str}")

# --- DROPDOWNS ---
print("\nData validations (dropdowns):")
for dv in ws.data_validations.dataValidation:
    sqref = " ".join(str(r) for r in dv.sqref.ranges) if hasattr(dv.sqref, "ranges") else str(dv.sqref)
    print(f"  - Range: {sqref} | Type: {dv.type} | Source: {dv.formula1}")

print("\nDIAGNOSIS COMPLETE")
