"""Build the HeyRoya Pro Correction Worksheet template (FINAL — last version).

Mirror of the latest `CreateHeyRoyaProTemplate_Final` macro.

Layout shifts vs. previous version:
  - Row 1: title "HeyRoya Correction Worksheet" (bold, size 14)
  - Rows 2-7: gray metadata box with thin borders, label/value split (A:B)
  - Row 8: italic "Numeric identifiers are quoted with a leading ' …" note
  - Row 10: italic instructions
  - Row 11: header row (9 columns, last_edited dropped)
  - Rows 12-500: data
  - Freeze panes at A12

Sheet protection: DISABLED per user request (no password). Cell `locked` flags
remain set so protection can be enabled later from the ribbon if desired.

Output: C:/Users/carin/Downloads/HeyRoya_Pro_Correction_Worksheet_Template.xlsx
        (or a -<timestamp> sibling if the base file is open in Excel)
"""

import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

_BASE = r"C:\Users\carin\Downloads\HeyRoya_Pro_Correction_Worksheet_Template.xlsx"

def _writable(path: str) -> str:
    if not os.path.exists(path):
        return path
    try:
        with open(path, "ab"):
            return path
    except PermissionError:
        ts = time.strftime("%Y%m%d-%H%M%S")
        root, ext = os.path.splitext(path)
        return f"{root}-{ts}{ext}"

OUT = _writable(_BASE)

GRAY    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
YELLOW  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN    = Side(style="thin", color="BFBFBF")
BOX     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD    = Font(bold=True)
TITLE   = Font(bold=True, size=14)
ITALIC  = Font(italic=True, color="555555")
SOFT    = Font(color="6A6A6A")

LAST_ROW = 500   # data rows 12..500

wb = Workbook()
ws = wb.active
ws.title = "HeyRoya_Correction"

# ---------------- 1. Title (row 1) ----------------
title_cell = ws.cell(row=1, column=1, value="HeyRoya Correction Worksheet")
title_cell.font = TITLE

# ---------------- 2. System metadata box (A2:F8 — includes IDs note row) ----------------
for r in range(2, 9):
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.fill = GRAY
        cell.border = BOX

ws.cell(row=2, column=1, value="SYSTEM-GENERATED — DO NOT EDIT").font = BOLD

ws.cell(row=3, column=1, value="Publisher:").font = BOLD
ws.cell(row=3, column=2, value="[ENTER NAME]")

ws.cell(row=4, column=1, value="Catalog:").font = BOLD
ws.cell(row=4, column=2, value="[ENTER CATALOG]")

ws.cell(row=5, column=1, value="Generated:").font = BOLD
# Pre-formatted string — Excel and any other viewer (preview, GPT, etc.)
# will display "2026-05-03 20:30" verbatim, no risk of seeing the raw
# serial number "46145.79…" because the cell type is text, not date.
gen_cell = ws.cell(
    row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"),
)

ws.cell(row=6, column=1, value="Completion:").font = BOLD
comp = ws.cell(
    row=6, column=2,
    value=f'=IFERROR(COUNTIF(E12:E{LAST_ROW},"<>")/COUNTIF(B12:B{LAST_ROW},"<>"),0)',
)
comp.number_format = "0%"

ws.cell(row=7, column=1, value="Status:").font = BOLD
ws.cell(
    row=7, column=2,
    value=(
        f'=IF(AND(COUNTIF(B12:B{LAST_ROW},"<>")>0,'
        f'COUNTIF(E12:E{LAST_ROW},"")=0,'
        f'COUNTIF(G12:G{LAST_ROW},"")=0),"READY","NOT READY")'
    ),
)

# ---------------- 3. Numeric IDs note (row 8) ----------------
ws.cell(row=8, column=1,
        value="Numeric identifiers are quoted with a leading ' for Excel compatibility."
        ).font = ITALIC

# ---------------- 4. Instructions (row 10) ----------------
ws.cell(row=10, column=1,
        value="Provide corrections in 'corrected_value'. System fields are locked."
        ).font = ITALIC

# ---------------- 5. Header row (row 11) — 9 cols A..I ----------------
headers = [
    "Line_ID", "work_id", "field", "original_value",
    "corrected_value", "decision_required",
    "status", "publisher_comment", "system_notes",
]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=11, column=col, value=h)
    c.font = BOLD
    c.fill = HEADER
    c.alignment = Alignment(horizontal="left")

# ---------------- 6. Data rows 12..LAST_ROW ----------------
for row in range(12, LAST_ROW + 1):
    # A: Line_ID — ROW()-11 produces 1, 2, 3 … on rows 12, 13, 14 …
    ws.cell(row=row, column=1, value=f'="L-" & TEXT(ROW()-11,"000")')
    # F: decision_required — blank on inactive rows
    ws.cell(row=row, column=6,
            value=f'=IF(B{row}="","",IF(E{row}="","Yes","No"))')

# Banding (A..I across the data range)
for row in range(12, LAST_ROW + 1):
    for col in (1, 2, 3, 4):
        ws.cell(row=row, column=col).fill = GRAY        # A..D system
    ws.cell(row=row, column=5).fill = YELLOW            # E corrected_value
    ws.cell(row=row, column=6).fill = GRAY              # F decision_required
    ws.cell(row=row, column=7).fill = YELLOW            # G status
    ws.cell(row=row, column=8).fill = YELLOW            # H publisher_comment
    ws.cell(row=row, column=9).fill = GRAY              # I system_notes

# ---------------- 7. Data validation ----------------
field_dv = DataValidation(
    type="list",
    formula1=(
        '"writer_ipi,writer_name,iswc,isrc,role_code,share,agreement_type,'
        'duplicate,identifier,title,work_title,publisher_ipi,writer_role"'
    ),
    allow_blank=True, showErrorMessage=True,
    errorTitle="Invalid field", error="Pick a value from the dropdown.",
)
ws.add_data_validation(field_dv)
field_dv.add(f"C12:C{LAST_ROW}")

status_dv = DataValidation(
    type="list",
    formula1='"Needs Review,Corrected,Confirmed Original,Not Applicable"',
    allow_blank=True, showErrorMessage=True,
    errorTitle="Invalid status", error="Pick a value from the dropdown.",
)
ws.add_data_validation(status_dv)
status_dv.add(f"G12:G{LAST_ROW}")

# ---------------- 8. Conditional formatting ----------------
# Green: corrected filled AND work_id present
ws.conditional_formatting.add(
    f"E12:E{LAST_ROW}",
    FormulaRule(
        formula=["AND(LEN(TRIM($E12))>0,LEN($B12)>0)"],
        stopIfTrue=False, fill=GREEN,
    ),
)
# Red: corrected non-empty AND original non-empty AND equal AND decision_required = Yes AND work_id present
ws.conditional_formatting.add(
    f"E12:E{LAST_ROW}",
    FormulaRule(
        formula=['AND($E12<>"",$D12<>"",$E12=$D12,$F12="Yes",LEN($B12)>0)'],
        stopIfTrue=False, fill=RED,
    ),
)
# Fade inactive rows (work_id blank) across A..I
ws.conditional_formatting.add(
    f"A12:I{LAST_ROW}",
    FormulaRule(
        formula=["LEN($B12)=0"],
        stopIfTrue=False, fill=GRAY,
    ),
)

# ---------------- 9. Freeze panes at A12 ----------------
ws.freeze_panes = "A12"

# ---------------- 10. Cell locking flags (sheet itself stays unprotected) ----------------
for row_iter in ws.iter_rows(min_row=1, max_row=LAST_ROW + 10, min_col=1, max_col=9):
    for cell in row_iter:
        cell.protection = Protection(locked=True)

for row in range(12, LAST_ROW + 1):
    ws.cell(row=row, column=5).protection = Protection(locked=False)  # corrected_value
    ws.cell(row=row, column=7).protection = Protection(locked=False)  # status
    ws.cell(row=row, column=8).protection = Protection(locked=False)  # publisher_comment

# ---------------- 11. Column widths ----------------
widths = {
    "A": 8,   # Line_ID
    "B": 12,  # work_id
    "C": 18,  # field
    "D": 28,  # original_value
    "E": 28,  # corrected_value
    "F": 14,  # decision_required
    "G": 16,  # status
    "H": 28,  # publisher_comment
    "I": 28,  # system_notes
}
for letter, w in widths.items():
    ws.column_dimensions[letter].width = w

# ---------------- 12. Footer (rows 505/506) ----------------
ws.cell(row=505, column=1,
        value=f"Worksheet generated by Kataloghub — {datetime.now():%Y-%m-%d %H:%M:%S}"
        ).font = SOFT
ws.cell(row=506, column=1,
        value="File-based correction workflow. No system access. All changes must be explicit."
        ).font = SOFT

# ---------------- 13. Sheet protection — DISABLED per user request (no password) ----------------
# ws.protection.sheet remains False; the lock flags above are dormant.

wb.save(OUT)

# ---------------- Post-process: bake cached values so non-Excel viewers
# display formula results without needing a calc engine. Excel itself will
# always re-evaluate the formula on open and overwrite these. ----------------
import re
import shutil
import zipfile

def bake_cached_values(xlsx_path: str) -> None:
    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                # B6 (Completion) — empty data → IFERROR returns 0 → 0% with the
                # 0% number format. Inject numeric cached value.
                text = re.sub(
                    r'(<c r="B6"[^>]*><f>[^<]+</f>)<v\s*/>',
                    r"\1<v>0</v>",
                    text, count=1,
                )
                # B7 (Status) — empty data → "NOT READY". Need t="str" on the
                # cell so the cached value is interpreted as inline string.
                text = re.sub(
                    r'(<c r="B7"\s)([^>]*?)(>)<f>([^<]+)</f><v\s*/>',
                    r'\1\2 t="str"\3<f>\4</f><v>NOT READY</v>',
                    text, count=1,
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, xlsx_path)

bake_cached_values(OUT)

print(f"OK -> {OUT}")
print("  cached values baked: B6=0, B7=NOT READY")
