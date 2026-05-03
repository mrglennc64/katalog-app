"""Build the HeyRoya Correction Worksheet template.

Mirror of the VBA macro `CreateHeyRoyaTemplate` — produces an .xlsx with
the same headers, gray locked / yellow editable bands, dropdowns on the
Field and Status columns, and the Carina-tone preamble + footer.
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\carin\Downloads\HeyRoya_Correction_Worksheet_Template.xlsx"

GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BOLD = Font(bold=True)
COMMENT_FONT = Font(color="6A6A6A", italic=True)

wb = Workbook()
ws = wb.active
ws.title = "HeyRoya Correction"

# 1. Header / Branding (Carina-tone)
preamble = [
    "# HeyRoya Correction Worksheet",
    "# File-based, zero-trust correction workflow",
    "# All corrections must be provided in the 'corrected_value' column.",
    "# Only edit yellow fields. Gray fields are system-generated.",
    "# Status dropdown: Needs Review / Corrected / Confirmed Original / Not Applicable",
]
for i, line in enumerate(preamble, start=1):
    cell = ws.cell(row=i, column=1, value=line)
    cell.font = COMMENT_FONT

ws.cell(row=7, column=1, value="# Publisher: ______________________").font = COMMENT_FONT
ws.cell(row=8, column=1, value="# Catalog: ________________________").font = COMMENT_FONT
ws.cell(row=9, column=1, value=f"# Date: {date.today().isoformat()}").font = COMMENT_FONT

# 2. Column headers
headers = ["work_id", "field", "original_value", "corrected_value", "status", "notes"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=11, column=col, value=h)
    c.font = BOLD
    c.alignment = Alignment(horizontal="left")

# 3. Zero-trust banding (rows 12..50)
for row in range(12, 51):
    for col in range(1, 4):  # A..C — system-generated, locked
        ws.cell(row=row, column=col).fill = GRAY
    for col in range(4, 7):  # D..F — publisher-editable
        ws.cell(row=row, column=col).fill = YELLOW

# 4. Data validation dropdowns
field_dv = DataValidation(
    type="list",
    formula1='"Writer Name,Writer IPI,ISWC,ISRC,Role Code,Share,Agreement,Identifier,Title,Publisher Name,Publisher IPI"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid field",
    error="Pick a value from the dropdown.",
)
ws.add_data_validation(field_dv)
field_dv.add("B12:B50")

status_dv = DataValidation(
    type="list",
    formula1='"Needs Review,Corrected,Confirmed Original,Not Applicable"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid status",
    error="Pick a value from the dropdown.",
)
ws.add_data_validation(status_dv)
status_dv.add("E12:E50")

# 5. Footer
footer = [
    (52, "# HeyRoya — Metadata Correction Service"),
    (53, "# This worksheet is part of the file-based correction pipeline."),
    (54, "# No system access. No ingestion. All changes must be explicit."),
]
for row, text in footer:
    ws.cell(row=row, column=1, value=text).font = COMMENT_FONT

# 6. Auto-fit (approximation: width = max content length per column)
for col in range(1, 7):
    letter = get_column_letter(col)
    max_len = 0
    for row in range(1, 55):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            max_len = max(max_len, len(str(v)))
    ws.column_dimensions[letter].width = max(12, min(60, max_len + 2))

wb.save(OUT)
print(f"OK -> {OUT}")
